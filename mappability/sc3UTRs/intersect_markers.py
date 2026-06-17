#!/usr/bin/env python3
"""
Intersect 3' window UF results with cell-type marker tables.

Loads all sc3prime_*.tsv files from results/ and crosses with:
  - CellMarker 2.0  (CellMarker2_Human_cell_markers.xlsx)
  - PanglaoDB       (PanglaoDB_markers_27Mar2020.tsv.gz)

Key question: which cell-type marker genes have uf_3p500 substantially below
uf_overall, making them unreliable in 10x Genomics 3' scRNA-seq?

Threshold: delta_3p500 < -0.10  (≥10% drop in 500 bp window)
           uf_3p500 < 0.50       (less than half of 3' reads map uniquely)

Outputs (in results/):
  marker_3prime_impact.tsv    — per marker gene × run: 3' UF metrics
  celltype_impact_summary.tsv — per cell type × database: n markers affected
  marker_flags.tsv            — per gene (any run): worst_delta, affected_runs
"""

import csv
import gzip
import os
import sys
from collections import defaultdict
from pathlib import Path

HERE    = Path(__file__).resolve().parent
RESULTS = HERE / 'results'

XLSX    = HERE / 'CellMarker2_Human_cell_markers.xlsx'
PANGL   = HERE / 'PanglaoDB_markers_27Mar2020.tsv'
FASTA   = Path(os.environ.get(
    'GENCODE_FASTA',
    HERE / '../../Ref/gencode.v49.transcripts.fa.gz'))
ANNOT   = RESULTS / 'gene_overlap_annotation.tsv'

DELTA_THR = -0.10   # UF drop threshold
UF_THR    = 0.50    # absolute UF threshold
WINDOW    = 'uf_3p500'
DELTA_COL = 'delta_3p500'


# ── ENST → gene_name mapping from GENCODE FASTA headers ──────────────────────

def build_enst_gene_map():
    """
    Parse GENCODE v49 transcript FASTA headers.
    Format: >ENST...|ENSG...|...|tx_name|gene_name|length|biotype|
    Returns {ENST_no_version: gene_name}
    """
    enst_gene = {}
    with gzip.open(FASTA, 'rt') as f:
        for line in f:
            if not line.startswith('>'):
                continue
            parts = line[1:].rstrip('|\n').split('|')
            enst  = parts[0].split('.')[0]
            gene  = parts[5] if len(parts) > 5 else ''
            if gene:
                enst_gene[enst] = gene
    print(f"ENST→gene map: {len(enst_gene):,} entries from GENCODE FASTA")
    return enst_gene


# ── gene_overlap_annotation → gene_name → phenomenon ─────────────────────────

def load_gene_annotation():
    """Returns {gene_name: {phenomenon, overlap_group_size, overlap_group_biotypes,
                             seq_cluster_id, seq_cluster_n_genes, seq_cluster_mixed}}"""
    annot = {}
    if not ANNOT.exists():
        print(f"WARNING: {ANNOT.name} not found — run annotate_gene_overlaps.py first")
        return annot
    with open(ANNOT) as f:
        for row in csv.DictReader(f, delimiter='\t'):
            gname = row['gene_name']
            if gname and gname not in annot:
                annot[gname] = {
                    'phenomenon':            row['phenomenon'],
                    'overlap_group_size':    row['overlap_group_size'],
                    'overlap_group_biotypes': row['overlap_group_biotypes'],
                    'seq_cluster_n_genes':   row['seq_cluster_n_genes'],
                    'seq_cluster_mixed':     row['seq_cluster_mixed'],
                }
    print(f"Gene annotation:  {len(annot):,} gene names loaded")
    return annot


# ── Load marker tables ────────────────────────────────────────────────────────

def load_cellmarker2():
    """Returns {gene_symbol: [(tissue_type, cell_name, technology), ...]}"""
    import openpyxl
    markers = defaultdict(list)
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb.active
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[col['species']] or row[col['species']] != 'Human':
            continue
        sym = row[col['Symbol']]
        if not sym:
            continue
        tissue = row[col.get('tissue_type', col.get('tissue_class', 0))] or ''
        cell   = row[col['cell_name']] or ''
        tech   = row[col.get('technology_seq', col.get('Genetype', 0))] or ''
        markers[str(sym).strip()].append((str(tissue), str(cell), str(tech)))
    print(f"CellMarker2:  {len(markers):,} unique marker genes")
    return markers


def load_panglao():
    """Returns {gene_symbol: [(organ, cell_type, species), ...]}"""
    markers = defaultdict(list)
    # File is stored as plain TSV (decompressed); gzip version handled transparently
    opener = gzip.open if str(PANGL).endswith('.gz') else open
    with opener(str(PANGL), 'rt') as fh:
        header = fh.readline().rstrip().split('\t')
        col = {h: i for i, h in enumerate(header)}
        for line in fh:
            parts = line.rstrip().split('\t')
            sp    = parts[col.get('species', 0)] if len(parts) > col.get('species', 0) else ''
            if 'Hs' not in sp:
                continue
            sym   = parts[col.get('official gene symbol', 1)].strip()
            organ = parts[col.get('organ', 9)].strip() if len(parts) > col.get('organ', 9) else ''
            ct    = parts[col.get('cell type', 2)].strip() if len(parts) > col.get('cell type', 2) else ''
            if sym:
                markers[sym].append((organ, ct, sp))
    print(f"PanglaoDB:    {len(markers):,} unique marker genes")
    return markers


# ── Load 3' UF results ────────────────────────────────────────────────────────

def load_3prime_results(enst_gene):
    """
    Returns {gene_symbol: {label: {col: value}}}.
    gene_symbol is looked up from enst_gene using transcript_id;
    the gene_symbol column in the TSV files is always empty.
    When multiple transcripts map to the same gene in one run,
    the one with the higher uf_overall is kept (MANE Select is usually
    the only entry per gene, but some genes have two MANE isoforms).
    """
    tsv_files = sorted(RESULTS.glob('sc3prime_*.tsv'))
    if not tsv_files:
        print(f"ERROR: no sc3prime_*.tsv files in {RESULTS}", file=sys.stderr)
        sys.exit(1)

    print(f"\nLoading {len(tsv_files)} 3\' UF result files ...", flush=True)
    gene_data  = defaultdict(dict)   # {gene_symbol: {label: row_dict}}
    all_labels = []

    for tsv in tsv_files:
        label = tsv.stem.replace('sc3prime_', '')
        all_labels.append(label)
        with open(tsv) as fh:
            for row in csv.DictReader(fh, delimiter='\t'):
                enst = row['transcript_id'].split('.')[0]
                gs   = enst_gene.get(enst, '').strip()
                if not gs:
                    continue
                row['gene_symbol'] = gs
                prev = gene_data[gs].get(label)
                if prev is None or float(row.get('uf_overall', 0)) > float(prev.get('uf_overall', 0)):
                    gene_data[gs][label] = row

    print(f"  Unique gene symbols: {len(gene_data):,}")
    print(f"  Run labels: {len(all_labels)}")
    return gene_data, sorted(all_labels)


# ── Summarise per gene across runs ────────────────────────────────────────────

def worst_metrics(gene_symbol, gene_data, all_labels):
    """For a gene, find worst delta_3p500 and which runs are affected."""
    rows = gene_data.get(gene_symbol, {})
    if not rows:
        return None, None, []

    worst_delta  = 0.0
    worst_uf     = 1.0
    affected     = []
    for label, row in rows.items():
        try:
            delta = float(row.get(DELTA_COL, 0))
            uf    = float(row.get(WINDOW, 1))
        except (ValueError, TypeError):
            continue
        if delta < worst_delta:
            worst_delta = delta
            worst_uf    = uf
        if delta < DELTA_THR or uf < UF_THR:
            affected.append(label)

    return worst_delta, worst_uf, affected


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    RESULTS.mkdir(parents=True, exist_ok=True)

    enst_gene  = build_enst_gene_map()
    gene_annot = load_gene_annotation()
    cm2        = load_cellmarker2()
    pangl      = load_panglao()
    gene_data, all_labels = load_3prime_results(enst_gene)

    all_markers = set(cm2) | set(pangl)
    print(f"\nTotal unique marker genes (both DBs): {len(all_markers):,}")

    # ── Output 1: per marker gene × run ──────────────────────────────────────
    impact_path = RESULTS / 'marker_3prime_impact.tsv'
    impact_fields = ['gene_symbol', 'in_cellmarker2', 'in_panglao', 'label',
                     'uf_overall', 'uf_3p250', 'uf_3p400', 'uf_3p500', 'uf_3p600',
                     'delta_3p250', 'delta_3p400', 'delta_3p500', 'delta_3p600',
                     'n_positions', 'transcript_len', 'read_len',
                     'phenomenon', 'overlap_group_size', 'overlap_group_biotypes',
                     'seq_cluster_n_genes', 'seq_cluster_mixed']
    impact_rows = []
    for sym in sorted(all_markers):
        in_cm2   = sym in cm2
        in_pangl = sym in pangl
        rows  = gene_data.get(sym, {})
        gannot = gene_annot.get(sym, {})
        for label in all_labels:
            row = rows.get(label)
            if not row:
                continue
            impact_rows.append({
                'gene_symbol':           sym,
                'in_cellmarker2':        int(in_cm2),
                'in_panglao':            int(in_pangl),
                'label':                 label,
                'uf_overall':            row.get('uf_overall', ''),
                'uf_3p250':              row.get('uf_3p250', ''),
                'uf_3p400':              row.get('uf_3p400', ''),
                'uf_3p500':              row.get('uf_3p500', ''),
                'uf_3p600':              row.get('uf_3p600', ''),
                'delta_3p250':           row.get('delta_3p250', ''),
                'delta_3p400':           row.get('delta_3p400', ''),
                'delta_3p500':           row.get('delta_3p500', ''),
                'delta_3p600':           row.get('delta_3p600', ''),
                'n_positions':           row.get('n_positions', ''),
                'transcript_len':        row.get('transcript_len', ''),
                'read_len':              row.get('read_len', ''),
                'phenomenon':            gannot.get('phenomenon', ''),
                'overlap_group_size':    gannot.get('overlap_group_size', ''),
                'overlap_group_biotypes': gannot.get('overlap_group_biotypes', ''),
                'seq_cluster_n_genes':   gannot.get('seq_cluster_n_genes', ''),
                'seq_cluster_mixed':     gannot.get('seq_cluster_mixed', ''),
            })

    with open(impact_path, 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=impact_fields, delimiter='\t')
        w.writeheader()
        w.writerows(impact_rows)
    print(f"\nMarker impact table ({len(impact_rows):,} rows) → {impact_path}")

    # ── Output 2: per-gene summary ────────────────────────────────────────────
    flags_path = RESULTS / 'marker_flags.tsv'
    flag_rows = []
    for sym in sorted(all_markers):
        worst_d, worst_uf, affected = worst_metrics(sym, gene_data, all_labels)
        in_uf_data = sym in gene_data
        in_cm2     = sym in cm2
        in_pangl   = sym in pangl
        gannot     = gene_annot.get(sym, {})

        cm2_cells   = ';'.join(sorted({c for _, c, _ in cm2.get(sym, [])}))
        pangl_cells = ';'.join(sorted({ct for _, ct, _ in pangl.get(sym, [])}))

        flag_rows.append({
            'gene_symbol':           sym,
            'in_cellmarker2':        int(in_cm2),
            'in_panglao':            int(in_pangl),
            'found_in_uf_data':      int(in_uf_data),
            'worst_delta_3p500':     round(worst_d, 4) if worst_d is not None else 'NA',
            'worst_uf_3p500':        round(worst_uf, 4) if worst_uf is not None else 'NA',
            'n_affected_runs':       len(affected),
            'affected_runs':         ';'.join(affected[:10]),
            'phenomenon':            gannot.get('phenomenon', ''),
            'overlap_group_size':    gannot.get('overlap_group_size', ''),
            'overlap_group_biotypes': gannot.get('overlap_group_biotypes', ''),
            'seq_cluster_n_genes':   gannot.get('seq_cluster_n_genes', ''),
            'seq_cluster_mixed':     gannot.get('seq_cluster_mixed', ''),
            'cellmarker2_cells':     cm2_cells[:500],
            'panglao_cells':         pangl_cells[:500],
        })

    flag_rows.sort(key=lambda r: r['worst_delta_3p500'] if r['worst_delta_3p500'] != 'NA' else 0)
    flags_fields = ['gene_symbol', 'in_cellmarker2', 'in_panglao', 'found_in_uf_data',
                    'worst_delta_3p500', 'worst_uf_3p500', 'n_affected_runs', 'affected_runs',
                    'phenomenon', 'overlap_group_size', 'overlap_group_biotypes',
                    'seq_cluster_n_genes', 'seq_cluster_mixed',
                    'cellmarker2_cells', 'panglao_cells']
    with open(flags_path, 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=flags_fields, delimiter='\t')
        w.writeheader()
        w.writerows(flag_rows)

    n_affected = sum(1 for r in flag_rows if r['n_affected_runs'] > 0)
    n_found    = sum(1 for r in flag_rows if r['found_in_uf_data'])
    print(f"Marker flags ({len(flag_rows):,} genes, {n_found:,} found in UF data,"
          f" {n_affected:,} affected) → {flags_path}")

    # ── Output 3: per-cell-type summary ──────────────────────────────────────
    celltype_path = RESULTS / 'celltype_impact_summary.tsv'
    ct_rows = []

    # CellMarker2: cell type = cell_name
    ct_genes = defaultdict(set)   # {(db, cell_type): {gene_symbol}}
    for sym, entries in cm2.items():
        for _, cell, _ in entries:
            if cell:
                ct_genes[('CellMarker2', cell)].add(sym)

    for sym, entries in pangl.items():
        for _, ct, _ in entries:
            if ct:
                ct_genes[('PanglaoDB', ct)].add(sym)

    affected_genes = {
        sym for r in flag_rows
        for sym in [r['gene_symbol']]
        if r['n_affected_runs'] > 0
    }

    for (db, cell), genes in sorted(ct_genes.items()):
        n_genes    = len(genes)
        n_in_uf    = sum(1 for g in genes if g in gene_data)
        n_affected = sum(1 for g in genes if g in affected_genes)
        ct_rows.append({
            'database':         db,
            'cell_type':        cell,
            'n_markers':        n_genes,
            'n_markers_in_uf':  n_in_uf,
            'n_markers_affected': n_affected,
            'frac_affected':    round(n_affected / n_in_uf, 4) if n_in_uf > 0 else 0.0,
            'affected_markers': ';'.join(sorted(g for g in genes if g in affected_genes))[:500],
        })

    ct_rows.sort(key=lambda r: -r['n_markers_affected'])
    ct_fields = ['database','cell_type','n_markers','n_markers_in_uf',
                 'n_markers_affected','frac_affected','affected_markers']
    with open(celltype_path, 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=ct_fields, delimiter='\t')
        w.writeheader()
        w.writerows(ct_rows)
    print(f"Cell type summary ({len(ct_rows):,} cell types) → {celltype_path}")

    # ── Quick summary to terminal ──────────────────────────────────────────────
    print(f"\n{'═'*60}")
    print("Markers with ≥1 affected run (delta_3p500 < {DELTA_THR} or uf < {UF_THR}):")
    print(f"  CellMarker2  : {sum(1 for r in flag_rows if r['in_cellmarker2'] and r['n_affected_runs'] > 0):,}")
    print(f"  PanglaoDB    : {sum(1 for r in flag_rows if r['in_panglao'] and r['n_affected_runs'] > 0):,}")
    print(f"{'─'*60}")
    print("Top 20 most-affected marker genes (worst delta_3p500):")
    for r in flag_rows[:20]:
        if r['worst_delta_3p500'] != 'NA' and float(r['worst_delta_3p500']) < DELTA_THR:
            print(f"  {r['gene_symbol']:<12}  delta={r['worst_delta_3p500']:>7}  "
                  f"uf={r['worst_uf_3p500']:>5}  runs={r['n_affected_runs']}")
    print(f"{'═'*60}")


if __name__ == '__main__':
    main()
